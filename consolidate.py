#!/usr/bin/python3

import sys
import bisect
import json
import openpyxl

if len(sys.argv) < 3:
    print('USAGE: {} <chart-of-accounts> <general-ledger>'.format(sys.argv[0]))
    sys.exit()
    
chart_path = sys.argv[1]
chart_wb = openpyxl.load_workbook(chart_path)
chart_sheet = chart_wb.active
chart_max_col = chart_sheet.max_column

ledger_path = sys.argv[2]
ledger_wb = openpyxl.load_workbook(ledger_path)
ledger_sheet = ledger_wb.active
ledger_max_col = ledger_sheet.max_column

if chart_max_col < 1:
    print("ERROR in chart accounts: too few columns!")
    sys.exit()
    
if ledger_max_col < 2:
    print("ERROR in general ledger: too few columns!")

chart_size = 0
ledger_size = 0
valid_strings = []
output = {}

while True:
    cell = chart_sheet.cell(row = chart_size+2, column=1)
    if cell.value == None:
        break
    chart_size += 1
    valid_strings.append(cell.value)
    
valid_strings.sort()

# Use this on a sorted list because binary search is O(log n)
# Using 'in' keyword in arbitrary list is O(n) 
def string_is_valid(word):
    idx = bisect.bisect(valid_strings, word)
    return valid_strings[idx-1:idx] == [word]
    
def increment_nested(data, args, num):
    if len(args) == 1:
        new_value = data.get("value", 0) + num
        data["value"] = new_value
    else:
        node = data.get(args[0],{})
        data[args[0]] = node
        increment_nested(node, args[1:], num)
        
while True:
    cellA = ledger_sheet.cell(row = ledger_size+2, column=1)
    if cellA.value == None:
        break
    cellB = ledger_sheet.cell(row = ledger_size+2, column=2)
    ledger_size += 1
    account = cellA.value
    value = cellB.value
    if not string_is_valid(account):
        print("ERROR: ledger contains invalid string")
        sys.exit()
    increment_nested(output,account.split('.'),value)
    
def adjust_sum(data):
    keys = list(data.keys())
    if len(data.keys()) == 1 and keys[0]=='value':
        return data['value']
    value = 0
    for str in data.keys():
        if (str != 'value'):
            value += adjust_sum(data[str])
    data['value'] = value
    return value

adjust_sum(output)            

print(json.dumps(output))

